"""
Processador de Acervo PJe — Streamlit
-------------------------------------
Recebe um xlsx exportado do PJe ("Relação do Acervo da Vara") e devolve um
relatório com:

  - aba "Relatório": dados originais + coluna ``data_mínima`` (= menor data
    entre ``dt_ultimo_movimento`` e ``data_entrada_tarefa``, apenas a data),
    organizados como Tabela do Excel ("Tabela1");
  - aba "tabela_dinamica": Tabela Dinâmica nativa do Excel, com
        Filtros : cargo_judicial, situacao_atual
        Linhas  : tarefa_atual
        Valores : Contagem de processo
        Colunas : anos (agrupado a partir de data_mínima).

A geração reaproveita um arquivo `template.xlsx` (que já contém a Tabela
Dinâmica configurada) — só substituímos os dados e marcamos o cache para
atualizar na abertura do Excel.

Nome do arquivo gerado: `relatorio_dinamico_dd_mm_aa_hh_mm_ss.xlsx`
"""

from __future__ import annotations

import sys
import unicodedata
import zipfile
from datetime import date, datetime
from io import BytesIO
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

try:
    import streamlit as st
except ImportError:  # pragma: no cover
    sys.exit("Streamlit não está instalado. Rode: pip install -r requirements.txt")

try:
    from streamlit.runtime.scriptrunner import get_script_run_ctx
    if get_script_run_ctx() is None:
        sys.exit(
            "Este app deve ser executado com:\n\n"
            "    streamlit run app.py\n\n"
            "Não use `python app.py` diretamente."
        )
except Exception:
    pass

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------
st.set_page_config(
    page_title="Relatório Dinâmico — Acervo PJe",
    page_icon="⚖️",
    layout="wide",
)

st.title("⚖️ Relatório Dinâmico — Acervo PJe")
st.caption(
    "Envie o xlsx exportado do PJe (Relação do Acervo da Vara, obtido em Relatórios Gerenciais - cód. 277). "
    "O app calcula data mínima, entre data da última movimentação e data de ingresso na tarefa, gerando um relatório com tabela dinâmica."
)

TEMPLATE_PATH = Path(__file__).parent / "template.xlsx"

HEADER_DETECT_COLUMNS = {
    "processo",
    "dt_ultimo_movimento",
    "data_entrada_tarefa",
    "cargo_judicial",
    "situacao_atual",
    "tarefa_atual",
}


# ---------------------------------------------------------------------------
# Utilidades
# ---------------------------------------------------------------------------
def _norm(s) -> str:
    """Remove acentos / casing / espaços para comparação de nomes de coluna."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )
    return s


def _detect_header(file_bytes: bytes) -> tuple[int, str]:
    wb = load_workbook(BytesIO(file_bytes), data_only=True, read_only=True)
    ws = wb.active
    sheet_name = ws.title
    header_row = None
    for r, row in enumerate(ws.iter_rows(values_only=True), start=1):
        values = {_norm(v) for v in row if v is not None}
        if HEADER_DETECT_COLUMNS.issubset(values):
            header_row = r
            break
        if r > 200:
            break
    wb.close()
    if header_row is None:
        raise ValueError(
            "Não foi possível localizar o cabeçalho. "
            "Verifique se o arquivo é a 'Relação do Acervo da Vara' do PJe - Relatórios Gerenciais 277."
        )
    return header_row, sheet_name


@st.cache_data(show_spinner=False)
def carregar(file_bytes: bytes) -> pd.DataFrame:
    """Lê o xlsx PJe e adiciona a coluna ``data_minima`` (apenas data)."""
    header_row, sheet_name = _detect_header(file_bytes)
    df = pd.read_excel(
        BytesIO(file_bytes),
        sheet_name=sheet_name,
        header=header_row - 1,
        engine="openpyxl",
    )
    df = df.dropna(how="all").reset_index(drop=True)

    dt_u = pd.to_datetime(df["dt_ultimo_movimento"], errors="coerce").dt.normalize()
    dt_e = pd.to_datetime(df["data_entrada_tarefa"], errors="coerce").dt.normalize()
    df["data_minima"] = pd.concat([dt_u, dt_e], axis=1).min(axis=1)
    return df


# ---------------------------------------------------------------------------
# Geração do relatório a partir do template
# ---------------------------------------------------------------------------
HEADER_ROW = 14  # linha do cabeçalho na aba "Relatório" do template


def _patch_pivot_refresh(xlsx_bytes: bytes) -> bytes:
    """Marca a Tabela Dinâmica para atualizar ao abrir o arquivo no Excel."""
    src = BytesIO(xlsx_bytes)
    dst = BytesIO()
    with zipfile.ZipFile(src) as zin, zipfile.ZipFile(
        dst, "w", zipfile.ZIP_DEFLATED
    ) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.startswith("xl/pivotCache/pivotCacheDefinition"):
                text = data.decode("utf-8")
                if "refreshOnLoad=" not in text:
                    text = text.replace(
                        "<pivotCacheDefinition ",
                        '<pivotCacheDefinition refreshOnLoad="1" ',
                        1,
                    )
                data = text.encode("utf-8")
            elif item.filename.startswith("xl/pivotTables/pivotTable"):
                text = data.decode("utf-8")
                # Garante atualização nas abas
                if 'updatedVersion="' in text and "refreshOnLoad=" not in text:
                    pass  # já tratado via cache
                data = text.encode("utf-8")
            zout.writestr(item, data)
    return dst.getvalue()


def gerar_xlsx(df_acervo: pd.DataFrame) -> bytes:
    """Substitui os dados do template e devolve os bytes do xlsx final."""
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(
            f"template.xlsx não encontrado em {TEMPLATE_PATH}. "
            "Coloque o template ao lado de app.py."
        )

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["Relatório"]

    # Cabeçalho do template
    template_cols = [
        ws.cell(HEADER_ROW, c).value for c in range(1, ws.max_column + 1)
    ]
    n_cols = len(template_cols)

    # Mapeia nome-normalizado do template → coluna real do DataFrame
    src_col_by_norm = {_norm(c): c for c in df_acervo.columns}
    # data_mínima do template ↔ data_minima do DataFrame
    src_col_by_norm.setdefault("data minima", "data_minima")

    # Limpa linhas de dados existentes no template
    existing_last = ws.max_row
    if existing_last > HEADER_ROW:
        for r in range(HEADER_ROW + 1, existing_last + 1):
            for c in range(1, n_cols + 1):
                ws.cell(r, c).value = None

    # Escreve novas linhas
    n_rows = len(df_acervo)
    for i, row in enumerate(df_acervo.itertuples(index=False), start=0):
        excel_row = HEADER_ROW + 1 + i
        row_dict = row._asdict()
        for col_idx, template_col in enumerate(template_cols, start=1):
            tnorm = _norm(template_col)
            src_name = src_col_by_norm.get(tnorm)
            if src_name is None or src_name not in row_dict:
                continue
            v = row_dict[src_name]
            if pd.isna(v):
                v = None
            elif isinstance(v, pd.Timestamp):
                v = v.to_pydatetime()
            ws.cell(excel_row, col_idx).value = v

    # Aplica formato de data curto em data_mínima
    try:
        dmin_idx = next(
            i for i, c in enumerate(template_cols, start=1) if _norm(c) == "data minima"
        )
        for r in range(HEADER_ROW + 1, HEADER_ROW + 1 + n_rows):
            ws.cell(r, dmin_idx).number_format = "DD/MM/YYYY"
    except StopIteration:
        pass

    # Atualiza o range da Tabela1 (e do autoFilter)
    last_row = HEADER_ROW + n_rows if n_rows else HEADER_ROW
    # Coluna final do template é AM (39)
    end_col_letter = "AM"
    new_ref = f"A{HEADER_ROW}:{end_col_letter}{last_row}"
    if "Tabela1" in ws.tables:
        tbl = ws.tables["Tabela1"]
        tbl.ref = new_ref
        if tbl.autoFilter is not None:
            tbl.autoFilter.ref = new_ref

    # Salva e marca o pivot para atualizar
    buf = BytesIO()
    wb.save(buf)
    return _patch_pivot_refresh(buf.getvalue())


# ---------------------------------------------------------------------------
# UI
# ---------------------------------------------------------------------------
uploaded = st.file_uploader(
    "Selecione o arquivo xlsx do acervo",
    type=["xlsx"],
    accept_multiple_files=False,
)

if not uploaded:
    st.info("Envie um arquivo para começar.")
    st.stop()

df: pd.DataFrame | None = None
try:
    df = carregar(uploaded.getvalue())
except Exception as exc:
    st.error(f"Erro ao processar o arquivo: {exc}")
    st.stop()

if df is None:
    st.stop()

st.success(f"Arquivo carregado: {len(df):,} linhas, {len(df.columns)} colunas.")

with st.expander("Pré-visualizar dados (primeiras 20 linhas)"):
    st.dataframe(df.head(20), use_container_width=True)

if st.button("📊 Gerar relatório xlsx", type="primary"):
    with st.spinner("Gerando arquivo..."):
        try:
            xlsx_bytes = gerar_xlsx(df)
        except Exception as exc:
            st.error(f"Erro ao gerar: {exc}")
            st.stop()

    ts = datetime.now().strftime("%d_%m_%y_%H_%M_%S")
    file_name = f"relatorio_dinamico_{ts}.xlsx"
    st.download_button(
        label=f"⬇️ Baixar {file_name}",
        data=xlsx_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    st.success(
        "Relatório pronto! Ao abrir no Excel, a Tabela Dinâmica é "
        "atualizada automaticamente."
    )
